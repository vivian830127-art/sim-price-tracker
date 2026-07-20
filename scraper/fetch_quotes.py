# -*- coding: utf-8 -*-
"""
自動登入 Fastmove 發貨系統 → 我的報價 → 下載報價 → 解析匯出檔
→ 產出 data/fastmove_quotes.json(加值方案 → 成本價)

帳密由環境變數 FM_USER / FM_PASS 提供(GitHub Secrets),程式碼中不含任何帳密。
使用 Playwright 模擬瀏覽器操作(填表單、按「登入」、按「下載報價」),
不依賴內部 API,後台介面小改版也不易失效。
"""
import json
import os
import re
import sys
import pathlib
import openpyxl
from playwright.sync_api import sync_playwright

ROOT = pathlib.Path(__file__).resolve().parent.parent
OUT = ROOT / "data" / "fastmove_quotes.json"
LOGIN_URL = "https://fmshippingsys.fastmove.com.tw/fmshippingsysadmin#/access/other/auth"
QUOTE_URL = "https://fmshippingsys.fastmove.com.tw/fmshippingsysadmin#/pages/quotemg/myquote"

norm = lambda s: re.sub(r"\s+", "", str(s or ""))


def download_report(user, pwd, save_to):
    with sync_playwright() as pw:
        browser = pw.chromium.launch()
        page = browser.new_page()
        page.goto(LOGIN_URL, wait_until="networkidle")

        # 登入表單:第一個文字框=帳號、第二個(password)=密碼
        page.locator("input").first.fill(user)
        page.locator("input[type=password]").first.fill(pwd)
        page.get_by_role("button", name="登入").click()
        page.wait_for_load_state("networkidle")

        page.goto(QUOTE_URL, wait_until="networkidle")
        with page.expect_download(timeout=60000) as dl:
            page.get_by_role("button", name="下載報價").click()
        dl.value.save_as(save_to)
        browser.close()


def parse_report(path):
    ws = openpyxl.load_workbook(path, data_only=True).active
    head = {str(ws.cell(1, c).value): c for c in range(1, ws.max_column + 1)}
    c_name, c_cost = head.get("商品名稱"), head.get("成本價(NT)")
    if not c_name or not c_cost:
        sys.exit(f"匯出檔欄位改變,找不到 商品名稱/成本價(NT):{list(head)}")
    quotes = {}
    for r in range(2, ws.max_row + 1):
        name, cost = ws.cell(r, c_name).value, ws.cell(r, c_cost).value
        if name and cost is not None:
            quotes[norm(name)] = float(cost)
    return quotes


def main():
    user, pwd = os.environ.get("FM_USER"), os.environ.get("FM_PASS")
    if not user or not pwd:
        sys.exit("缺少 FM_USER / FM_PASS 環境變數(請到 repo Settings → Secrets 設定)")
    tmp = ROOT / "data" / "_fm_report.xlsx"
    tmp.parent.mkdir(parents=True, exist_ok=True)
    download_report(user, pwd, tmp)
    quotes = parse_report(tmp)
    tmp.unlink(missing_ok=True)  # 用完即刪,不留在 repo
    OUT.write_text(json.dumps(quotes, ensure_ascii=False, indent=1), encoding="utf-8")
    print(f"取得 Fastmove 報價 {len(quotes)} 筆 → {OUT.name}")


if __name__ == "__main__":
    main()
