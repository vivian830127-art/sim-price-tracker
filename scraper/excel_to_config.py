# -*- coding: utf-8 -*-
"""
讀取 config/PChome售價試算.xlsx → 產出 config/products.json
- 完整保留 25 個欄位所需的原始資料
- 合併儲存格(加值方案/APN標籤)自動向下填補
- 從「參考連結」解析 PChome 商品 ID,供 fetch_prices.py 抓即時售價
計算公式(已與 Excel 逐列驗算一致):
  單件成本   = 浮動報價 × 天數 × 1.05
  平台前毛/後扣/物流費 = 依「平台抽成設定」表查表
  平台進價   = ROUNDDOWN(售價 - 售價×前毛)      (蝦皮無此欄)
  銷售成本   = 蝦皮: 售價×(前毛+後扣)+物流費
               其他: 售價×前毛 + 進價×後扣 + 物流費
  單件毛利   = ROUNDDOWN(售價 - 銷售成本)
  總毛利     = 單件毛利 × 商品數量
  單品利潤   = 單件毛利 - 單件成本
  總利潤     = 總毛利 - 總成本價
"""
import json
import math
import re
import pathlib
import openpyxl

ROOT = pathlib.Path(__file__).resolve().parent.parent
XLSX = ROOT / "config" / "PChome售價試算.xlsx"
OUT = ROOT / "config" / "products.json"
SHEET = "PC叫貨表1"
FIRST_ROW, LAST_COL = 4, 25

COLS = {  # 欄位 → Excel 欄號
    "商品ID": 1, "料號": 2, "標籤產品名稱": 3, "加值方案": 4, "APN標籤": 5,
    "天數": 6, "國家": 7, "用量": 8, "參考連結": 9, "平台": 10,
    "浮動報價": 11, "售價": 13, "商品數量": 14, "網路類型": 25,
}


def read_fee_table(ws):
    fees = {}
    for r in range(3, 20):
        name = ws.cell(r, 26).value
        if name:
            fees[str(name).strip()] = {
                "前毛": ws.cell(r, 27).value or 0,
                "後扣": ws.cell(r, 28).value or 0,
                "物流費": ws.cell(r, 29).value or 0,
            }
    return fees


def fill_merged(ws):
    """把合併儲存格的值填到範圍內每一格(加值方案、APN 等)。"""
    filled = {}
    for rng in ws.merged_cells.ranges:
        v = ws.cell(rng.min_row, rng.min_col).value
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                filled[(r, c)] = v
    return filled


def cell(ws, filled, r, c):
    return filled.get((r, c), ws.cell(r, c).value)


def compute(price, qty, quote, days, plat, fees):
    f = fees.get(plat, {"前毛": 0, "後扣": 0, "物流費": 0})
    fm, hk, lg = f["前毛"] or 0, f["後扣"] or 0, f["物流費"] or 0
    unit_cost = round((quote or 0) * (days or 0) * 1.05, 2)
    total_cost = round(unit_cost * qty, 2)
    if plat == "蝦皮":
        buy_in = None
        sell_cost = price * (fm + hk) + lg
    else:
        buy_in = math.floor(price - price * fm)
        sell_cost = price * fm + buy_in * hk + lg
    sell_cost = round(sell_cost, 2)
    unit_gross = math.floor(price - sell_cost)
    total_gross = unit_gross * qty
    unit_profit = round(unit_gross - unit_cost, 2)
    total_profit = round(total_gross - total_cost, 2)
    return {
        "單件成本": unit_cost, "總成本價": total_cost,
        "平台前毛": fm, "平台進價": buy_in, "平台後扣": hk, "物流空卡費": lg,
        "銷售成本": sell_cost, "單件毛利": unit_gross, "總毛利": total_gross,
        "單品利潤": unit_profit, "總利潤": total_profit,
    }


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb[SHEET]
    fees = read_fee_table(ws)
    filled = fill_merged(ws)

    products = []
    for r in range(FIRST_ROW, ws.max_row + 1):
        price = cell(ws, filled, r, COLS["售價"])
        name = cell(ws, filled, r, COLS["標籤產品名稱"])
        country = cell(ws, filled, r, COLS["國家"])
        if price is None or country is None:
            continue  # 空白列或合計列

        link = str(cell(ws, filled, r, COLS["參考連結"]) or "").strip()
        m = re.search(r"/prod/([A-Z0-9\-]+)", link)
        pchome_id = m.group(1) if m else None

        raw = {k: cell(ws, filled, r, c) for k, c in COLS.items()}
        qty = raw["商品數量"] or 1
        quote = raw["浮動報價"] or 0
        days = raw["天數"] or 0
        plat = str(raw["平台"] or "PChome").strip()

        key = raw["商品ID"] or pchome_id or f"ROW{r}"
        products.append({
            "key": str(key),
            "商品ID": raw["商品ID"] or "",
            "料號": str(int(raw["料號"])) if isinstance(raw["料號"], float) else (raw["料號"] or ""),
            "標籤產品名稱": name or f"{country}{int(days)}天(未命名)",
            "加值方案": (raw["加值方案"] or "").strip() if raw["加值方案"] else "",
            "APN標籤": raw["APN標籤"] or "",
            "天數": int(days), "國家": country, "用量": raw["用量"] or "",
            "參考連結": link if link.startswith("http") else "",
            "參考說明": "" if link.startswith("http") else link,
            "pchome_id": pchome_id,
            "平台": plat, "網路類型": raw["網路類型"] or "",
            "浮動報價": quote, "售價": price, "商品數量": qty,
            **compute(price, qty, quote, days, plat, fees),
        })

    OUT.write_text(json.dumps({"fees": fees, "products": products},
                              ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"轉換完成:{len(products)} 個品項 → {OUT.name}")


if __name__ == "__main__":
    main()
